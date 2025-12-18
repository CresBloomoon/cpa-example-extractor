import re
import json
import zipfile
from dataclasses import dataclass, asdict
from typing import List, Optional, Tuple, Dict
from io import BytesIO

import fitz  # PyMuPDF
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# =========================
# 科目ラベル（表示は日本語で統一）
# =========================

SUBJECT_LABELS = {
    "zeimu": "租税",
    "zaimu": "財務",
    "kanri": "管理",
    "unknown": "不明",
}
LABEL_TO_CODE = {v: k for k, v in SUBJECT_LABELS.items()}

SUBJECT_CODES = ["zeimu", "zaimu", "kanri", "unknown"]
SUBJECT_LABEL_OPTIONS = [SUBJECT_LABELS[c] for c in SUBJECT_CODES]


# =========================
# 正規化ユーティリティ
# =========================

def normalize_dashes(s: str) -> str:
    return (
        s.replace("－", "-")
         .replace("―", "-")
         .replace("−", "-")
    )

def normalize_rank(rank: Optional[str]) -> Optional[str]:
    if not rank:
        return None
    return rank.translate(str.maketrans({"Ａ": "A", "Ｂ": "B", "Ｃ": "C"}))


# =========================
# 正規表現（共通：まずはこれで走らせる）
# =========================

EXAMPLE_RE = re.compile(
    r"例題\s*(?P<num>\d+)\s+(?P<title>.*?)(?:\s*（(?P<rank>[A-CＡ-Ｃ])）)?\s*(?=\n|$)"
)

PAGE_REF_RE = re.compile(
    r"法(?P<section>[^\s\-]+)-(?P<pageno>\d+)"
)


# =========================
# 科目判定：キーワードスコア方式（堅い）
# =========================

SUBJECT_KEYWORDS: Dict[str, Dict[str, int]] = {
    "zeimu": {  # 租税
        "法人税": 6, "法人税法": 6, "租税": 5, "租税公課": 6, "税効果会計": 4,
        "申告": 3, "課税所得": 4, "別表": 4, "受取配当": 3, "完全支配関係": 3,
        "寄附金": 3, "交際費": 3, "減価償却": 3,
    },
    "zaimu": {  # 財務会計
        "財務会計": 6, "連結": 5, "企業結合": 5, "金融商品": 4, "退職給付": 4,
        "包括利益": 3, "キャッシュ・フロー": 3, "会計方針": 3, "減損": 3,
        "収益認識": 3, "資産除去債務": 3, "リース": 3,
    },
    "kanri": {  # 管理会計
        "管理会計": 6, "CVP": 6, "標準原価": 5, "差異分析": 5, "直接原価": 5,
        "予算": 4, "意思決定": 4, "設備投資": 4, "原価計算": 4, "原価企画": 4,
        "部門別": 3, "内部振替": 3,
    },
}

# ファイル名は強いヒントなので、少し強めに効かせる
FILENAME_WEIGHT = 1.8

# ファイル名にありがちな表記ゆれも拾う用（加点キーワード）
FILENAME_HINTS: Dict[str, Dict[str, int]] = {
    "zeimu": {
        "租税": 6, "税法": 6, "法人税": 6, "法人税法": 6, "消費税": 6, "所得税": 6,
    },
    "zaimu": {
        "財務": 6, "財務会計": 6, "会計基準": 4, "連結": 4, "企業結合": 4,
    },
    "kanri": {
        "管理": 6, "管理会計": 6, "原価": 4, "CVP": 6, "差異": 4, "予算": 4,
    },
}


def _score_text(text: str, weights: Dict[str, Dict[str, int]], base_factor: float = 1.0) -> Dict[str, int]:
    scores = {k: 0 for k in ["zeimu", "zaimu", "kanri"]}
    t = text or ""
    for subj, kws in weights.items():
        s = 0
        for kw, w in kws.items():
            if kw and kw in t:
                s += w
        scores[subj] += int(round(s * base_factor))
    return scores


def detect_subject_from_doc(doc: fitz.Document, file_name: str) -> Tuple[str, Dict[str, int]]:
    # 1) TOC
    try:
        toc_text = " ".join([t for _, t, _ in doc.get_toc() if t]) or ""
    except Exception:
        toc_text = ""

    # 2) 先頭ページ本文（軽く）
    head_text = ""
    for i in range(min(8, len(doc))):
        try:
            head_text += (doc[i].get_text("text") or "") + "\n"
        except Exception:
            pass

    toc_text = normalize_dashes(toc_text).replace("\u00a0", " ")
    head_text = normalize_dashes(head_text).replace("\u00a0", " ")

    # 3) ファイル名（強いヒント）
    fname = normalize_dashes(file_name or "").replace("\u00a0", " ")

    # スコア合算
    scores = {k: 0 for k in ["zeimu", "zaimu", "kanri"]}

    s1 = _score_text(toc_text, SUBJECT_KEYWORDS, base_factor=1.0)
    s2 = _score_text(head_text, SUBJECT_KEYWORDS, base_factor=1.0)
    s3 = _score_text(fname, FILENAME_HINTS, base_factor=FILENAME_WEIGHT)

    for k in scores:
        scores[k] = s1[k] + s2[k] + s3[k]

    # best
    best = max(scores, key=scores.get)
    ordered = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    top_score = ordered[0][1]
    second_score = ordered[1][1]

    # 安全運転：弱い/拮抗なら不明
    if top_score < 6 or (top_score - second_score) < 2:
        return "unknown", scores

    return best, scores


# =========================
# データ構造：章/節（アウトライン由来）
# =========================

@dataclass
class Chapter:
    no: int
    title: str
    start_page: int
    sections: List["Section"]

@dataclass
class Section:
    no: int
    title: str
    start_page: int

@dataclass
class ExampleItem:
    subject: str
    chapter_no: int
    chapter_title: str
    section_no: int
    section_title: str
    example_no: int
    title: str
    rank: Optional[str]
    page_ref: Optional[str]
    pdf_page: int
    source_pdf: str


def parse_outline(doc: fitz.Document) -> List[Chapter]:
    chapters: List[Chapter] = []
    current_chapter: Optional[Chapter] = None

    for level, title, page in doc.get_toc():
        title = normalize_dashes(title or "")

        if level == 1:
            m = re.search(r"第(\d+)章\s*(.+)", title)
            if not m:
                continue
            current_chapter = Chapter(
                no=int(m.group(1)),
                title=m.group(2),
                start_page=page,
                sections=[]
            )
            chapters.append(current_chapter)

        elif level == 2 and current_chapter:
            m = re.search(r"第(\d+)節\s*(.+)", title)
            if not m:
                continue
            current_chapter.sections.append(
                Section(
                    no=int(m.group(1)),
                    title=m.group(2),
                    start_page=page
                )
            )

    return chapters


def find_chapter_section(chapters: List[Chapter], pdf_page: int) -> Tuple[Optional[Chapter], Optional[Section]]:
    current_ch = None
    for ch in chapters:
        if pdf_page >= ch.start_page:
            current_ch = ch
    if not current_ch:
        return None, None

    current_sec = None
    for sec in current_ch.sections:
        if pdf_page >= sec.start_page:
            current_sec = sec

    return current_ch, current_sec


def extract_page_ref(text: str) -> Optional[str]:
    m = PAGE_REF_RE.search(text)
    if not m:
        return None
    return f"法{m.group('section')}-{m.group('pageno')}"


# =========================
# 抽出（共通）：科目ごとに差し替え可能な設計
# =========================

def extract_examples_common(pdf_bytes: bytes, subject_code: str, source_pdf: str) -> List[ExampleItem]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    chapters = parse_outline(doc)

    results: List[ExampleItem] = []
    last_page_ref: Optional[str] = None

    for i in range(len(doc)):
        page_no = i + 1
        page = doc[i]
        text = normalize_dashes(page.get_text("text") or "").replace("\u00a0", " ")

        found_ref = extract_page_ref(text)
        page_ref = found_ref or last_page_ref
        if found_ref:
            last_page_ref = found_ref

        chapter, section = find_chapter_section(chapters, page_no)
        if not chapter or not section:
            continue

        for m in EXAMPLE_RE.finditer(text):
            raw_title = re.sub(r"\s+", " ", (m.group("title") or "").strip())
            title = re.sub(r"\s*（[A-CＡ-Ｃ]）\s*$", "", raw_title)

            results.append(
                ExampleItem(
                    subject=subject_code,
                    chapter_no=chapter.no,
                    chapter_title=chapter.title,
                    section_no=section.no,
                    section_title=section.title,
                    example_no=int(m.group("num")),
                    title=title,
                    rank=normalize_rank(m.group("rank")),
                    page_ref=page_ref,
                    pdf_page=page_no,
                    source_pdf=source_pdf,
                )
            )

    doc.close()
    return results


def extract_examples(pdf_bytes: bytes, subject_code: str, source_pdf: str) -> List[ExampleItem]:
    # 今は共通抽出器。科目別に変えたい時はここで分岐する
    return extract_examples_common(pdf_bytes, subject_code, source_pdf)


# =========================
# ソート（アップロード順に依存しない）
# =========================

def sort_df(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["chapter_no", "section_no", "example_no", "pdf_page"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    sort_cols = [c for c in ["subject", "chapter_no", "section_no", "pdf_page", "example_no"] if c in df.columns]
    return df.sort_values(by=sort_cols, kind="stable").reset_index(drop=True)


# =========================
# None 集計（全セル）
# =========================

def count_none_cells(df: pd.DataFrame) -> Tuple[int, Dict[str, int]]:
    na_mask = df.isna()
    none_str_mask = df.applymap(lambda x: isinstance(x, str) and x.strip().lower() == "none")
    mask = na_mask | none_str_mask
    total = int(mask.to_numpy().sum())
    per_col = mask.sum().astype(int).to_dict()
    return total, {k: int(v) for k, v in per_col.items()}


# =========================
# Excel 出力（科目単位）
#  - 章ごとにシート
#  - シート内で節区切り
#  - 例題行：A=連番, B=title, C=rank, D=page_ref（B/C入替）
# =========================

def build_excel(df: pd.DataFrame, subject_code: str) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    df_sorted = sort_df(df.copy())

    if "subject" in df_sorted.columns:
        df_sorted = df_sorted[df_sorted["subject"] == subject_code]

    if df_sorted.empty:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    for (chap_no, chap_title), chap_df in df_sorted.groupby(["chapter_no", "chapter_title"], sort=True):
        chap_no_int = int(chap_no) if pd.notna(chap_no) else 0
        base_name = f"{chap_no_int}章 {chap_title}"[:31]
        sheet_name = base_name

        idx = 2
        while sheet_name in used_names:
            suffix = f"_{idx}"
            sheet_name = (base_name[:31 - len(suffix)] + suffix)
            idx += 1

        used_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        row = 1
        for (sec_no, sec_title), sec_df in chap_df.groupby(["section_no", "section_title"], sort=True):
            sec_no_int = int(sec_no) if pd.notna(sec_no) else 0

            ws.cell(row=row, column=1, value=f"第{sec_no_int}節")
            ws.cell(row=row, column=2, value=sec_title)
            row += 1

            counter = 1
            for _, r in sec_df.iterrows():
                ws.cell(row=row, column=1, value=counter)
                ws.cell(row=row, column=2, value=r.get("title"))     # B: title
                ws.cell(row=row, column=3, value=r.get("rank"))      # C: rank
                ws.cell(row=row, column=4, value=r.get("page_ref"))  # D: page_ref
                counter += 1
                row += 1

            row += 1

        widths = {1: 10, 2: 46, 3: 10, 4: 14}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =========================
# ZIP 出力（科目別xlsx + json）
# =========================

def build_zip(per_subject_dfs: Dict[str, pd.DataFrame]) -> BytesIO:
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for subj in ["zeimu", "zaimu", "kanri"]:
            df = per_subject_dfs.get(subj)
            if df is None or df.empty:
                continue

            label = SUBJECT_LABELS[subj]
            xlsx_buf = build_excel(df, subj)
            zf.writestr(f"{label}_examples.xlsx", xlsx_buf.getvalue())

            j = json.dumps(df.to_dict(orient="records"), ensure_ascii=False, indent=2)
            zf.writestr(f"{label}_examples.json", j)

        unk = per_subject_dfs.get("unknown")
        if unk is not None and not unk.empty:
            j = json.dumps(unk.to_dict(orient="records"), ensure_ascii=False, indent=2)
            zf.writestr("不明_examples.json", j)

        all_df = pd.concat([d for d in per_subject_dfs.values() if d is not None], ignore_index=True)
        all_df = sort_df(all_df)
        zf.writestr("ALL_examples.json", json.dumps(all_df.to_dict(orient="records"), ensure_ascii=False, indent=2))

    zip_buf.seek(0)
    return zip_buf


# =========================
# Streamlit UI
# =========================

st.set_page_config(page_title="CPA 例題抽出（科目自動判定→科目別出力）", layout="wide")
st.title("CPAテキスト 例題抽出ツール（科目自動判定 → 不明のみ手修正 → 科目別xlsx）")

uploaded_files = st.file_uploader(
    "PDFをドラッグ＆ドロップ（複数OK）",
    type=["pdf"],
    accept_multiple_files=True
)

if uploaded_files:
    file_rows = []
    file_bytes_map: Dict[str, bytes] = {}

    with st.spinner("科目判定中…（ファイル名も加点）"):
        for uf in uploaded_files:
            b = uf.read()
            file_bytes_map[uf.name] = b

            doc = fitz.open(stream=b, filetype="pdf")
            subj_code, scores = detect_subject_from_doc(doc, uf.name)
            doc.close()

            file_rows.append({
                "file_name": uf.name,
                "detected_subject": SUBJECT_LABELS[subj_code],
                "score_租税": scores.get("zeimu", 0),
                "score_財務": scores.get("zaimu", 0),
                "score_管理": scores.get("kanri", 0),
                "final_subject": SUBJECT_LABELS[subj_code],
            })

    st.success(f"アップロード：{len(uploaded_files)} ファイル")

    file_df = pd.DataFrame(file_rows)

    st.subheader("科目判定結果（不明だけ直せばOK）")

    with st.form("subject_form", clear_on_submit=False):
        edited_file_df = st.data_editor(
            file_df,
            use_container_width=True,
            num_rows="fixed",
            column_config={
                "final_subject": st.column_config.SelectboxColumn(
                    "final_subject",
                    options=SUBJECT_LABEL_OPTIONS
                ),
                "detected_subject": st.column_config.TextColumn("detected_subject", disabled=True),
                "file_name": st.column_config.TextColumn("file_name", disabled=True),
                "score_租税": st.column_config.NumberColumn("score_租税", disabled=True),
                "score_財務": st.column_config.NumberColumn("score_財務", disabled=True),
                "score_管理": st.column_config.NumberColumn("score_管理", disabled=True),
            },
            key="subject_editor"
        )
        run_extract = st.form_submit_button("解析実行（ここで抽出スタート）")

    if run_extract:
        all_items = []
        with st.spinner("例題抽出中…（少し待ってね）"):
            for _, r in edited_file_df.iterrows():
                fname = r["file_name"]
                subj_label = r["final_subject"]
                subj_code = LABEL_TO_CODE.get(subj_label, "unknown")

                b = file_bytes_map[fname]
                items = extract_examples(b, subj_code, fname)
                all_items.extend([asdict(x) for x in items])

        if not all_items:
            st.warning("例題が抽出できませんでした。例題の表記ゆれがあるかも。")
            st.stop()

        base_df = sort_df(pd.DataFrame(all_items))
        st.session_state["base_df"] = base_df
        st.session_state["edited_df"] = base_df.copy()

    if "base_df" in st.session_state:
        st.divider()
        st.subheader("編集（最終チェック）")

        with st.form("edit_form", clear_on_submit=False):
            edited_df = st.data_editor(
                st.session_state["edited_df"],
                use_container_width=True,
                num_rows="fixed",
                column_config={
                    "rank": st.column_config.SelectboxColumn("rank", options=[None, "A", "B", "C"]),
                    "title": st.column_config.TextColumn("title"),
                    "source_pdf": st.column_config.TextColumn("source_pdf", disabled=True),
                },
                key="main_editor"
            )
            update_check = st.form_submit_button("チェック更新（ここで集計）")

        if update_check:
            st.session_state["edited_df"] = sort_df(edited_df)

        current_df = st.session_state["edited_df"]

        st.subheader("チェック結果（現在）")
        total_none, per_col = count_none_cells(current_df)
        st.metric("全セルの未入力（None）件数", f"{total_none} 個")

        st.caption("列別の未入力（None）件数")
        per_col_df = (
            pd.DataFrame([{"column": k, "none_count": int(v)} for k, v in per_col.items()])
            .sort_values("none_count", ascending=False)
            .reset_index(drop=True)
        )
        st.dataframe(per_col_df, use_container_width=True)

        if total_none == 0:
            st.success("🎉 未入力はありません。出力してOKです。")

        st.divider()
        st.subheader("最終出力")

        per_subject = {}
        for s in SUBJECT_CODES:
            per_subject[s] = current_df[current_df["subject"] == s].copy() if "subject" in current_df.columns else pd.DataFrame()

        zip_buf = build_zip(per_subject)

        st.download_button(
            "科目別ZIPをダウンロード（xlsx + json）",
            data=zip_buf,
            file_name="CPA_examples_by_subject.zip",
            mime="application/zip"
        )

        st.download_button(
            "ALL_examples.json（まとめ）をダウンロード",
            data=json.dumps(sort_df(current_df).to_dict(orient="records"), ensure_ascii=False, indent=2),
            file_name="ALL_examples.json",
            mime="application/json"
        )
