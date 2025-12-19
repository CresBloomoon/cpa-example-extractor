# cpa_tool/excel_export.py
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def _safe_sheet_title(title: str) -> str:
    """
    Excelのシート名制約対応（31文字 / 禁止文字）
    """
    if title is None:
        title = ""
    bad = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in bad:
        title = title.replace(ch, " ")
    title = title.strip()
    if not title:
        title = "Sheet"
    return title[:31]


def _write_header(ws, headers: List[str]) -> None:
    ws.append(headers)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autofit_columns(ws, max_width: int = 60) -> None:
    # 雑に列幅調整（超安全寄り）
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, min(ws.max_row, 300) + 1):  # コスト抑える
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)


def build_excel(df: pd.DataFrame, subj: str) -> BytesIO:
    """
    df: 1科目分の抽出結果
    subj: 'zaimu' / 'kanri' / 'zeimu' など
    """
    wb = Workbook()

    # ★ 重要：必ず1枚は残す（openpyxlの例外回避）
    ws_default = wb.active
    ws_default.title = "INFO"
    ws_default["A1"] = "No data"
    ws_default["A2"] = f"subject={subj}"
    ws_default["A3"] = "抽出結果が0件でした。PDF / 判定 / 抽出条件を確認してください。"
    ws_default["A1"].font = Font(bold=True)

    if df is None or len(df) == 0:
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # NaNをNoneに寄せる
    df = df.copy()
    df = df.where(pd.notnull(df), None)

    # よく使う列（無ければ空でOK）
    # ※ 以前の要望で「B列とC列入れ替え」とかあったけど、
    #    ここでは “データとしての列” をそのまま出す（UI側で並べ替えしてもOK）
    cols = list(df.columns)

    # 章単位でシートを切る（chapter_no, chapter_title が無い場合は1枚にまとめる）
    has_chapter = ("chapter_no" in df.columns) or ("chapter_title" in df.columns)

    if has_chapter:
        # chapter_no でソート（文字でも落ちないように安全に）
        def _chapter_sort_key(x):
            try:
                return int(x) if x is not None else 10**9
            except Exception:
                return 10**9

        # グループキー作成
        df["_chapter_no_sort"] = df.get("chapter_no", None).apply(_chapter_sort_key)
        df = df.sort_values(by=["_chapter_no_sort", "chapter_title", "section_no", "example_no"], na_position="last")
        df = df.drop(columns=["_chapter_no_sort"], errors="ignore")

        # groupby（chapter_no/title が片方無くても動く）
        group_keys = []
        if "chapter_no" in df.columns:
            group_keys.append("chapter_no")
        if "chapter_title" in df.columns:
            group_keys.append("chapter_title")

        grouped = df.groupby(group_keys, dropna=False) if group_keys else [(("ALL",), df)]

        created_any = False
        for key, g in grouped:
            # key は tuple か scalar
            if not isinstance(key, tuple):
                key = (key,)
            chap_no = None
            chap_title = None
            if "chapter_no" in group_keys:
                chap_no = key[group_keys.index("chapter_no")]
            if "chapter_title" in group_keys:
                chap_title = key[group_keys.index("chapter_title")]

            sheet_name = ""
            if chap_no is not None:
                sheet_name += f"{chap_no}章 "
            if chap_title:
                sheet_name += str(chap_title)
            sheet_name = _safe_sheet_title(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            created_any = True

            _write_header(ws, cols)

            for _, row in g.iterrows():
                ws.append([row.get(c, None) for c in cols])

            _autofit_columns(ws)

        # 1枚でも作れたら、INFOは消してOK
        if created_any and "INFO" in wb.sheetnames:
            wb.remove(wb["INFO"])

    else:
        # 章情報が無いなら、全部まとめて1枚
        ws = wb.create_sheet(title=_safe_sheet_title(subj))
        _write_header(ws, cols)
        for _, row in df.iterrows():
            ws.append([row.get(c, None) for c in cols])
        _autofit_columns(ws)

        # INFO消す
        if "INFO" in wb.sheetnames:
            wb.remove(wb["INFO"])

    # 念押し：見えるシートが無い状態を絶対作らない
    if len(wb.worksheets) == 0:
        ws = wb.create_sheet(title="INFO")
        ws["A1"] = "No data (safety sheet)"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
